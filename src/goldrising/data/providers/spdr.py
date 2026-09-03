"""SPDR Gold Trust 官方历史档案（xlsx）：GLD 收盘、NAV、持金吨数。一手来源。"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

import pandas as pd

from goldrising.contracts import Indicator, IndicatorSeries, Provenance
from goldrising.data.providers.base import FetchContext, ProviderError

URL = "https://api.spdrgoldshares.com/api/v1/historical-archive"
COL_CLOSE = "Closing Price"
COL_OZ_PER_SHARE = "Ounces of Gold per Share"
COL_NAV = "NAV/Share at 10:30am NYT"
COL_TONNES = "Tonnes of Gold"


def load_gld_archive(ctx: FetchContext) -> pd.DataFrame:
    key = "spdr:gld_archive"
    cached = ctx.cache.get(key)
    if isinstance(cached, pd.DataFrame):
        return cached
    import openpyxl  # 函数级导入

    resp = ctx.get(URL, params={"product": "gld", "exchange": "NYSE", "lang": "en"}, raw_name="spdr/gld_archive.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    target = None
    for name in wb.sheetnames:
        ws = wb[name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first and first[0] == "Date":
            target = ws
            break
    if target is None:
        raise ProviderError("SPDR 档案中找不到以 Date 开头的数据表")
    rows = target.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows)]
    records: list[dict[str, Any]] = []
    for r in rows:
        if not r or r[0] is None:
            continue
        try:
            d = datetime.strptime(str(r[0]), "%d-%b-%Y").date()
        except ValueError:
            continue
        rec: dict[str, Any] = {"date": d}
        for h, v in zip(header[1:], r[1:], strict=False):
            rec[h] = v if isinstance(v, int | float) else None
        records.append(rec)
    if not records:
        raise ProviderError("SPDR 档案无有效行")
    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"])
    df = df.set_index("date").sort_index()
    ctx.cache[key] = df
    return df


class SpdrProvider:
    name = "spdr"

    def fetch(self, indicator: Indicator, ctx: FetchContext, since: date | None = None) -> IndicatorSeries:
        df = load_gld_archive(ctx)
        series = indicator.source.series
        if series == "close":
            s = pd.to_numeric(df[COL_CLOSE], errors="coerce")
            note = "GLD 收盘价"
        elif series == "tonnes":
            s = pd.to_numeric(df[COL_TONNES], errors="coerce")
            note = "信托持金吨数"
        elif series == "nav_implied_gold":
            nav = pd.to_numeric(df[COL_NAV], errors="coerce")
            oz = pd.to_numeric(df[COL_OZ_PER_SHARE], errors="coerce")
            s = nav / oz
            note = "NAV/份 ÷ 每份含金盎司；LBMA 定盘价已被 ICE IBA 要求从档案中移除，此为其官方 NAV 口径的等价推算"
        else:
            raise ProviderError(f"{indicator.id}: 未知 SPDR 序列 {series}")
        s = s.dropna()
        prov = Provenance(
            provider=self.name,
            publisher=indicator.source.publisher or "SPDR Gold Trust / State Street Global Advisors",
            series=f"US GLD Historical Archive / {series}",
            url=indicator.source.url,
            fetched_at=Provenance.now_iso(),
            as_of=s.index[-1].strftime("%Y-%m-%d"),
            lag_days=indicator.lag_days,
            note=note,
        )
        return IndicatorSeries(indicator.id, s, prov)
